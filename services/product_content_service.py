"Product Content Service"

import os
from openpyxl.utils import get_column_letter
from models.account_credentials import AccountCredentials
from models.excel_cells import ExcelCells
from models.excel_file import ExcelFile
from models.web_driver import WebDriver
from utils.excel_utils import get_column_data_from_excel_sheet, open_workbook
from utils.file_utils import save_file

from utils.hepsiburada_utils import (
    get_product_info_from_hepsiburada,
    get_product_url_from_hepsiburada,
    handle_hepsiburada_cookie,
    login_to_hepsiburada,
)
from utils.jinja_utils import create_html_from_jinja_template
from utils.omniens_utils import get_product_info_from_omniens, login_to_omniens
from utils.trendyol_utils import (
    get_product_info_from_trendyol,
    get_product_url_from_trendyol,
    login_to_trendyol,
)

# pylint: disable=too-many-locals
# pylint: disable=too-many-arguments


class ProductContentService:
    "Product Content Service Class"

    def get_product_codes_from_excel(
        self,
        excel_file: ExcelFile,
        product_code_cells: ExcelCells,
    ):
        "Gets all product codes from specified column and in excel sheet"

        excel_file.workbook = open_workbook(excel_file.file_path)
        excel_file.sheet = excel_file.workbook[excel_file.sheet_name]

        product_codes = get_column_data_from_excel_sheet(
            excel_file.sheet,
            product_code_cells.column_start,
            product_code_cells.row_start,
            product_code_cells.row_end,
        )

        excel_file.workbook.close()

        return [str(code) for code in product_codes]

    def save_trendyol_product_urls_to_excel(
        self,
        browser: WebDriver,
        trendyol_credentials: AccountCredentials,
        excel_file: ExcelFile,
        trendyol_product_url_cells: ExcelCells,
        product_codes: [str],
    ) -> dict:
        "Gets the product URLs from Trendyol and saves them to excel file"

        excel_file.workbook = open_workbook(excel_file.file_path)
        excel_file.sheet = excel_file.workbook[excel_file.sheet_name]
        product_urls = {}

        login_to_trendyol(
            browser,
            trendyol_credentials.username,
            trendyol_credentials.password,
        )
        for i, product_code in enumerate(product_codes):
            product_url = get_product_url_from_trendyol(browser, product_code)
            product_urls[product_code] = product_url
            cell_column = get_column_letter(trendyol_product_url_cells.column_start)
            cell_row = trendyol_product_url_cells.row_start + i
            if product_url:
                excel_file.sheet[f"{cell_column}{cell_row}"] = product_url
                print(f"{cell_row} - {product_code} - Trendyol URL: {product_url}")
            else:
                excel_file.sheet[f"{cell_column}{cell_row}"] = ""
                print(f"{cell_row} - {product_code} - Not Found On Trendyol")
            excel_file.workbook.save(excel_file.file_path)
        excel_file.workbook.close()

        return product_urls

    def save_hepsiburada_product_urls_to_excel(
        self,
        browser: WebDriver,
        hepsiburada_credentials: AccountCredentials,
        excel_file: ExcelFile,
        hepsiburada_product_url_cells: ExcelCells,
        product_codes: [str],
    ) -> dict:
        "Gets the product URLs from Hepsiburada and saves them to excel file"

        excel_file.workbook = open_workbook(excel_file.file_path)
        excel_file.sheet = excel_file.workbook[excel_file.sheet_name]
        product_urls = {}

        login_to_hepsiburada(
            browser,
            hepsiburada_credentials.username,
            hepsiburada_credentials.password,
        )
        for i, product_code in enumerate(product_codes):
            product_url = get_product_url_from_hepsiburada(browser, product_code)
            product_urls[product_code] = product_url
            cell_column = get_column_letter(hepsiburada_product_url_cells.column_start)
            cell_row = hepsiburada_product_url_cells.row_start + i
            if product_url:
                excel_file.sheet[f"{cell_column}{cell_row}"] = product_url
                print(f"{cell_row} - {product_code} - Hepsiburada URL: {product_url}")
            else:
                excel_file.sheet[f"{cell_column}{cell_row}"] = ""
                print(f"{cell_row} - {product_code} - Not Found On Hepsiburada")
            excel_file.workbook.save(excel_file.file_path)
        excel_file.workbook.close()

        return product_urls

    def compare_product_info(
        self,
        browser: WebDriver,
        omniens_credentials: AccountCredentials,
        excel_file: ExcelFile,
        trendyol_product_name_cells: ExcelCells,
        hepsiburada_product_name_cells: ExcelCells,
        product_codes: [str],
        trendyol_urls: dict,
        hepsiburada_urls: dict,
    ):
        "Compares the Trendyol and Hepsiburada product informations with Omniens"

        excel_file.workbook = open_workbook(excel_file.file_path)
        excel_file.sheet = excel_file.workbook[excel_file.sheet_name]
        products_with_all_desc = []

        login_to_omniens(
            browser,
            omniens_credentials.username,
            omniens_credentials.password,
        )
        for i, product_code in enumerate(product_codes):
            trendyol_product_name, trendyol_product_desc = None, None
            hepsiburada_product_name, hepsiburada_product_desc = None, None
            omniens_product_name, omniens_product_desc = get_product_info_from_omniens(
                browser, product_code
            )
            if product_code in trendyol_urls.keys():
                (
                    trendyol_product_name,
                    trendyol_product_desc,
                ) = get_product_info_from_trendyol(
                    browser,
                    trendyol_urls[product_code],
                )
                cell_column = get_column_letter(
                    trendyol_product_name_cells.column_start
                )
                cell_row = trendyol_product_name_cells.row_start + i
                excel_file.sheet[f"{cell_column}{cell_row}"] = (
                    "DOĞRU"
                    if trendyol_product_name == omniens_product_name
                    else "YANLIŞ"
                )
                excel_file.workbook.save(excel_file.file_path)
            if product_code in hepsiburada_urls.keys():
                handle_hepsiburada_cookie(browser)
                (
                    hepsiburada_product_name,
                    hepsiburada_product_desc,
                ) = get_product_info_from_hepsiburada(
                    browser,
                    hepsiburada_urls[product_code],
                )
                cell_column = get_column_letter(
                    hepsiburada_product_name_cells.column_start
                )
                cell_row = hepsiburada_product_name_cells.row_start + i
                excel_file.sheet[f"{cell_column}{cell_row}"] = (
                    "DOĞRU"
                    if hepsiburada_product_name == omniens_product_name
                    else "YANLIŞ"
                )
                excel_file.workbook.save(excel_file.file_path)
            if trendyol_product_desc or hepsiburada_product_desc:
                products_with_all_desc.append(
                    {
                        "code": product_code,
                        "omniens_desc": omniens_product_desc,
                        "trendyol_desc": trendyol_product_desc,
                        "hepsiburada_desc": hepsiburada_product_desc,
                    }
                )
                comparison_html = create_html_from_jinja_template(
                    os.path.join("templates", "jinja"),
                    os.path.join("product_desc_comparison.html"),
                    {"products": products_with_all_desc},
                )
                save_file(
                    comparison_html,
                    os.path.join("temp", "product_description_comparison.html"),
                )
        excel_file.workbook.close()
